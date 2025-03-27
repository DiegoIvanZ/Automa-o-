from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os 


def search (a2):
    wb = load_workbook('Dados.xlsx')
    ws = wb.active

    for y in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if tuple(y[:3]) == tuple(a2):  # Comparar a linha inteira com a entrada
            messagebox.showinfo(title="Existente", message="Já existente!")
            return
    messagebox.showinfo(title="Non existence", message="Inexistente")


def remotion(a3):
    wb = load_workbook('Dados.xlsx')
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_col=3, values_only=False):
        if tuple(cell.value for cell in row[:3]) == tuple(a3):  # Comparar a linha com os dados fornecidos
            print('Dados encontrados. Removendo linha...')
            ws.delete_rows(row[0].row)  # Deleta a linha
            wb.save('Dados.xlsx')
            messagebox.showinfo(title="Remoção", message="Cadastro removido com sucesso!")
            return  # Para sair da função após encontrar e remover

