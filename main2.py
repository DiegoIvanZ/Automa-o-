from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os 

def cadastro(a1):
    if not os.path.exists('Dados.xlsx'):
        # Criar o arquivo se não existir
        wb = Workbook()
        ws = wb.active
        ws.append(['Nome', 'Email', 'Telefone'])  # Cabeçalhos
    else:
        # Abrir o arquivo existente
        wb = load_workbook('Dados.xlsx')
        ws = wb.active
    existing_data = set()
    
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
         existing_data.add(tuple(row))

        

    if tuple(a1) in existing_data:
        messagebox.showinfo(title="Error", message="Dados Duplicados")
        return 

    ws.append([a1[0], a1[1], a1[2]])
    wb.save('Dados.xlsx')
    messagebox.showinfo(title="Cadastro", message="Cadastro executado!.")
    